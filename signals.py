import pandas as pd
import json
import re
import io
import os

from schemas import SignalParams
from fastapi import UploadFile
from typing import Union

from openpyxl.styles import Border, Side, PatternFill


class Signals:
    def __init__(self, json_file_path: str):
        self.data = self.load_data(json_file_path)

    @staticmethod
    def separate_params(param: str):
        pattern = r"([a-zA-Z]+)(\d+)"

        matches = re.match(pattern, param)

        if matches:
            char = matches.group(1).upper()
            index = int(matches.group(2))

            valid_chars = ['X', 'M', 'Y', 'V', 'CV', 'T']
            if char not in valid_chars:
                raise ValueError(f"Invalid parameter character: {char}")

            match char:
                case 'X' | 'Y' | 'T':
                    if index < 0 or index > 1023:
                        raise ValueError(f"Invalid index for parameter {char}: {index}")
                case 'M':
                    if index < 0 or index > 12287:
                        raise ValueError(f"Invalid index for parameter {char}: {index}")
                case 'V':
                    if index < 0 or index > 14847:
                        raise ValueError(f"Invalid index for parameter {char}: {index}")
                case 'CV':
                    if index < 0 or index > 255:
                        raise ValueError(f"Invalid index for parameter {char}: {index}")

            return char, index
        else:
            raise ValueError("Invalid parameter format")

    @staticmethod
    def load_data(json_file_path: str):
        try:
            with open(json_file_path, 'r') as json_file:
                return json.load(json_file)
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {json_file_path}")
        except json.JSONDecodeError:
            raise ValueError(f"Invalid JSON format in file: {json_file_path}")

    def get_signals_by_param(self, param: str):
        char, index = self.separate_params(param)
        if not isinstance(index, int):
            raise ValueError(f"Invalid id: {index}. It should be an integer.")

        data_for_char = self.data.get(char, None)
        if data_for_char:
            if index < len(data_for_char):
                return data_for_char[index]
            else:
                raise IndexError(f"Index {index} out of range for parameter: {char}")
        else:
            raise ValueError(f"No data found for parameter: {char}-{index}")

    def get_signals_by_model(self, signal_params: SignalParams):
        params = signal_params.to_sequence_list()
        return self.get_signals_by_params(params[1:])

    def get_signals_by_params(self, params: list[str]):
        return [{
            'signal': param.upper(),
            'result': self.get_signals_by_param(param)
        } for param in params if param]

    def get_signals_by_xlsx_params(self, xlsx_file: Union[UploadFile, bytes]):
        if isinstance(xlsx_file, UploadFile):
            file = xlsx_file.file
        else:
            file = io.BytesIO(xlsx_file)

        excel_data = pd.read_excel(file)

        variables = excel_data['Переменная']

        dec_values = [self.get_signals_by_param(variable)['dec'] for variable in variables]

        excel_data['DEC'] = dec_values

        return excel_data

    @staticmethod
    def save_data_to_file(processed_data: pd.DataFrame, file_name: str):
        path_to_directory = "static/preprocessed_data"
        path_to_file = f"{path_to_directory}/{file_name}"

        if not os.path.exists(path_to_directory):
            os.makedirs(path_to_directory)

        with pd.ExcelWriter(path_to_file, engine='openpyxl') as writer:
            processed_data.to_excel(writer, sheet_name='Sheet1', index=False)
            worksheet = writer.sheets['Sheet1']

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = Border(
                        left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin')
                    )
                    row_index = cell.row
                    if row_index == 1:
                        cell.fill = PatternFill(start_color="4fad5c", end_color="4fad5c", fill_type="solid")

            for col_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in col_cells)
                worksheet.column_dimensions[col_cells[0].column_letter].width = length + 2

        return path_to_file
